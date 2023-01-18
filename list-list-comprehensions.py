#  [n**2 for n in numbers if n%2==0 if n % 6]

# F-strings

# class BankClass:
#     def __init__(self):
#         pass

#     amount = 0

# class SavingAccount(BankClass):
    
#     def __init__(self):
#         super().__init__()
    
#     def increase(self, increaseAmount):
#         self.amount += increaseAmount
#         print(f"Increased {str(increaseAmount)}")
#         print(f"New amount is {str(self.amount)}")
    
#     def decrease(self, decreaseAmount):
#         self.amount -= decreaseAmount
#         print(f"decreased {str(decreaseAmount)}")
#         print(f"New amount is {str(self.amount)}")

# account1 = SavingAccount()
# account1.increase(100)
# account1.decrease(50)

# class Shoe:
#     def __init__(self):
#         self.size = 0
#         self.color = "null"
#         self.model = "null"
#         self.waterproof = False
#     def info(self):
#         print(f"Size {self.size} -- Color {self.color} -- Model {self.model} -- Waterproof {self.waterproof}")

# class Sandal(Shoe):
#     def __init__(self):
#         super().__init__()
        

# class Converse(Shoe):
#     def __init__(self):
#         super().__init__()

# class Boot(Shoe):
#     def __init__(self):
#         super().__init__()
#         self.waterproof = True

# sandal1 = Boot()
# sandal1.size = 40
# sandal1.color = "Blue"
# sandal1.model = "A5435H"
# sandal1.info()

# class Address:
#     def __init__(self):
#         self.street = "Pastevrova"
#         self.number = 15

# class CampusAddress(Address):
#     def __init__(self, officeNumber):
#         super().__init__()
#         self.officeNumber = officeNumber
#     def info(self):
#         print(f"Street: {self.street}, Number: {str(self.number)}, Office Number: {str(self.officeNumber)}")
# office1 = CampusAddress(5.15)
# office1.info()

# numbers = range(1000)
# list = [n for n in numbers if n % 7 == 0]
# print(list)

# numbers = range(1000)
# list = [n for n in numbers if "3" in str(n)]
# print(list)

# letterArr = list("asdas asd a asd asds asd asd asd asd asd assd asd asd     ")
# print(letterArr)
# list1 = [lett for lett in letterArr if lett == " " ]
# # print(len(list1))
# letterArr = list("Yellow yaks like yelling and yawking and yesturday they yodled while eating yuky yams")
# novel = ["a", "e", "i", "o", "ö", "ü", "u", "ı"]
# newList = [lett for lett in letterArr if not lett.lower() in novel]

# print("".join(newList))

# tuple1 = ("hi", 4, 8.99, "apple", ('t,b','n'))

# for item in tuple1:
#     print(f"({tuple1.index(item)}, {item})")

# list_a = [1, 2, 3, 4]
# list_b = [2, 3, 4, 5]
# common_list = []
# for item_a in list_a:
#     for item_b in list_b:
#         if(item_a == item_b):
#             common_list.append(item_b)
# print(common_list)

# numbers = (0,1,2)
# x,y,z, = numbers

# customer = {
#     "name" : "Berkay Ceylan",
#     "age" : 26,
#     "is_verified" : True
# }

# customer["bithday"] = "July 15 1996"

# print(customer["bithday"])
# print(customer.get("birthday", "July 15 1996"))

# phone = input("Phone: ")
# digits_mapping = {
#     "0" : "Zero",
#     "1" : "One",
#     "2" : "Two",
#     "3" : "Three",
#     "4" : "Four",
#     "5" : "Five",
#     "6" : "Six",
#     "7" : "Seven",
#     "8" : "Eight",
#     "9" : "Nine",
# }
# output = ""

# for ch in phone:
#     output += digits_mapping.get(ch, "!") + " "
# print(output)

# message = input("write a message >")
# words = message.split(" ")
# print(words)
# emojis = {
#     ":)" : "😄",
#     ":(" : "🙁"
# }
# output = ""

# for word in words:
#     output += emojis.get(word, word) + " "
# print(output)

# numbers = [5,2,1,7,4]
# numbers.append(20)
# print(numbers)
# numbers.insert(0,10)
# print(numbers)

# numbers.remove(5)
# print(numbers)

# numbers.pop()
# print(numbers)
# print(numbers.index(2))

# numbers.clear()
# print(numbers)

# numbers = [5,2,1,5,7,4]
# print(numbers.index(5))

# numbers.sort()
# print(numbers)

# numbers.reverse()
# print(numbers)

# numbers2 = numbers.copy()
# numbers2.append(10)
# print(numbers2)


# dic = {
#     "0": 50,
#     "1": 20,
#     "2": 70,
#     "3": 10
# }
# dicSorted = sorted(dic.items(), key=lambda x:x[1])
# print(dicSorted)

dic1 = {1: 10, 2: 20}
dic2 = {3: 30, 4: 40}
dic3 = {5: 50, 6: 60}

# dicMerge = dic1.copy()
# dicMerge.update(dic2)
# dicMerge.update(dic3)

# # print(dicMerge)

# def checkValue(dic, value):
#     available = False
#     for item in dic.keys():
#         if value == item:
#             available = True
#     return available
# customer = {
#     "name" : "Berkay Ceylan",
#     "age" : 26,
#     "is_verified" : True
# }
# print(checkValue(customer, "namesdas"))
# dic1 = {1: 10, 2: 20}
# dic2 = {3: 30, 4: 40}
# dic3 = {5: 50, 6: 60}

# merged = {}

# for item in dic1.keys():
#     merged[item] = dic1[item]

# for item in dic2.keys():
#     merged[item] = dic2[item]

# for item in dic3.keys():
#     merged[item] = dic3[item]

# print(merged)

# def powerNumb(count):
#     finalDic = {}
#     for numb in range(count):
#         numb += 1
#         finalDic[numb] = numb*numb
#     return finalDic

# print(powerNumb(5))

import random

# for i in range(3):
#     print(random.randint(10, 20))

# members = ["Marie", "John", "Petr", "Lukas"]
# leader = random.choice(members)
# print(leader)

# class Dice:
#     def roll(self):
#         first = random.randint(1,6)
#         second = random.randint(1,6)
#         return(first, second)

# dice = Dice()
# print(dice.roll())

# from pathlib import Path

# path = Path()

# for file in path.glob("*.py"):
#     print(file)

import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook("Book1.xlsx")
sheet = wb["Sheet1"]
cell = sheet["a1"]
cell = sheet.cell(1,1)
print(cell.value)
print(sheet.max_row)

for row in range(1, sheet.max_row +1):
    print(row)

for row in range(2, sheet.max_row +1):
    cell = sheet.cell(row, 3)
    corrected_price = float(cell.value) * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

values = Reference(sheet, min_row = 2, max_row = sheet.max_row, min_col = 4, max_col = 4)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')

wb.save("new.xlsx")